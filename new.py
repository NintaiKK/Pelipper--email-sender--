from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from tkinter import messagebox
import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.ttk import *
from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import time
import pyautogui

print(r'''
 _____        _  _
|  __ \      | |(_)                                 
| |__) |  __ | | _  _ __   _ __    ___  _ __      _______
|  ___/ / _ \| || || '_ \ | '_ \  / _ \| '__|    |==   []|   
| |    |  __/| || || |_) || |_) ||  __/| |       |  ==== |  
|_|     \___||_||_|| .__/ | .__/  \___||_|       '-------'        
                   | |    | |
                   |_|    |_|

        1.0.2 beta            by PakecitusPaki
''')

janela = tk.Tk()
janela.resizable(0,0)
janela.geometry('800x500')
janela.title('Pelipper -Whatsapp bulk sender-')

a = tk.Frame()
b = tk.Frame()
bb = tk.Frame()
c = tk.Frame()

selected_data = []

def opcoes():
    global selected_data

    wb = load_workbook('pelipper_db.xlsx')
    ws = wb.active

    janelaOpc = tk.Tk()
    janelaOpc.title('Contatos')

    frame = tk.Frame(janelaOpc)
    frame.pack(fill = tk.BOTH, expand = True)

    colunas = [cell.value for cell in ws[1]]
    tree = ttk.Treeview(frame, columns =colunas, show = 'headings', selectmode = 'none')
    tree.pack(fill=tk.BOTH, expand = True)

    for col in colunas:
        tree.heading(col, text = col)
        tree.column(col, width = 100, anchor = 'w')

    for row in ws.iter_rows(min_row = 2, values_only = True):
        row = tuple('' if value is None else value for value in row)
        tree.insert("", tk.END, values = row)

    selected_itens = []

    def atualizar_destaque():
        for item in tree.get_children():
            tree.item(item, tags= '')
        for item in selected_itens:
            tree.item(item, tags = 'selected')

    def selecionar_linha(event):
        item = tree.identify_row(event.y)
        if item:
            if item in selected_itens:
                selected_itens.remove(item)
            else:
                selected_itens.append(item)
            atualizar_destaque()

    def selecionar_todas_linhas():
        global selected_data
        selected_itens[:] = [item for item in tree.get_children()]  # Atualiza a lista de itens selecionados
        atualizar_destaque()

    def confirmar_selecao():
        global selected_data
        selected_data = [tree.item(item, 'values') for item in selected_itens]
        janelaOpc.destroy()

    tree.tag_configure('selected', background='lightblue')

    tree.bind("<ButtonRelease-1>", selecionar_linha)

    select_all_btn = tk.Button(janelaOpc, text="Selecionar Todas", command=selecionar_todas_linhas)
    select_all_btn.pack(pady=10)

    confirm_btn = tk.Button(janelaOpc, text="Confirmar Seleção", command=confirmar_selecao)
    confirm_btn.pack(pady=10)

    janelaOpc.mainloop()

def trabaio():
    
    global selected_data

    login = enLogin.get()
    senha = enSenha.get()

    assunto = assunt.get('1.0', 'end')
    print(assunto)

    mensagem = msg.get('1.0', 'end')
    print(mensagem)

    workbook = load_workbook('pelipper_db.xlsx')
    sheet = workbook.active
    dados = [{'email': row[0], 'numero': row[1]} for row in sheet.iter_rows(min_row=2, values_only=True)]

    servico = Service(ChromeDriverManager().install())
    navegador = webdriver.Chrome(service=servico)

    navegador.get("https://mail.hostinger.com")

    try:
        xLogin = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="rcmloginuser"]'))
        )
        xLogin.click()
        xLogin.send_keys(login)
    except Exception as e:
        messagebox.showerror('Erro', 'O campo "usuário" não foi carregada corretamente, tente novamente mais tarde')
        break

    try:
        xSenha = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="rcmloginpwd"]'))
        )
        xSenha.click()
        xSenha.send_keys(senha)
    except Exception as e:
        messagebox.showerror('Erro', 'O campo "senha" não foi carregada corretamente, tente novamente mais tarde')
        break

    try:
        xEntrar = WebDriverWait(navegador, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="rcmloginsubmit"]'))
        )
        xEntrar.click()
    except Exception as e:
        messagebox.showerror('Erro', 'O botão "entrar" não foi carregada corretamente, tente novamente mais tarde')
        break


    for item in selected_data:
        email, numero = item

        try:
            xCriar = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="rcmbtn101"]'))
            )
            xCriar.click()
        except Exception as e:
            messagebox.showerror('Erro', 'O botão "criar nova mensagem" não foi carregada corretamente, tente novamente mais tarde')
            break
            
        try:
            xEndereço = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="compose_to"]/div/div/ul/li/input'))
            )
            xEndereço.click()
            xEndereço.send_keys(email)
        except Exception as e:
            messagebox.showerror('Erro', 'O campo "para" não foi carregada corretamente, tente novamente mais tarde')
            break

        try:
            xAssunto = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="compose-subject"]'))
            )
            xAssunto.click()
            xAssunto.send_keys(assunto)
        except Exception as e:
            messagebox.showerror('Erro', 'O campo "assunto" não foi carregada corretamente, tente novamente mais tarde')
            break


        try:
            xMensagem = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="composebodycontainer"]/div[1]/div[1]/div[2]'))
            )
            xMensagem.click()
            xMensagem.send_keys(mensagem)
        except Exception as e:
            messagebox.showerror('Erro', 'O campo "mensagem" não foi carregada corretamente, tente novamente mais tarde')
            break

        try:
            xEnviar = WebDriverWait(navegador, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="rcmbtn110"]'))
            )
            xEnviar.click()
        except Exception as e:
            messagebox.showerror('Erro', 'O campo "mensagem" não foi carregada corretamente, tente novamente mais tarde')
            break

        print(f'{empresa} enviado')

lblVrd = tk.Label(
    master= a,
    height = 30,
    width = 35,
    text='Pelipper v1.0.2',
    bg='light blue')
lblVrd.pack()

lbl_coiso = tk.Label(
    master = a,
    width=35,
    text='Licenciado para:',
    bg='light blue')
lbl_coiso.pack()

lbl_coisos = tk.Label(
    master = a,
    width=35,
    text='Protection Assessoria Empresarial®',
    bg='light blue')
lbl_coisos.pack()

lblSpcs = tk.Label(
    master = c,
    width = 20)
lblSpcs.pack()

###
#adicionar entrys
#login email
#senha email
#assunto
###

lblDI = tk.Label(
    master = b,
    text = 'Mensagem')
lblDI.pack()

lblSpcss = tk.Label(
    master = bb,
    width = 1)
lblSpcss.pack()

msg = tk.Text(
    master = b,
    height = 3,
    width = 30)
msg.pack()

fff = tk.Label(
    master = b,
    height = 3)
fff.pack()

opc = tk.Button(
    master = b,
    width = 10,
    text = 'Contatos',
    command = opcoes)
opc.pack()

ddd = tk.Label(
    master = b,
    height = 1)
ddd.pack()

ok = tk.Button(
    master = b,
    height = 1,
    text = 'Enviar',
    command = trabaio)
ok.pack()

a.pack(side = LEFT)
c.pack(side = LEFT)
b.pack(side = LEFT)
bb.pack(side = LEFT)


janela.mainloop()
